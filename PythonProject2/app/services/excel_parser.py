import pandas as pd
from datetime import datetime, timedelta
from typing import List, Dict, Any, Optional
import logging
import openpyxl
from openpyxl import load_workbook

logger = logging.getLogger(__name__)


class ExcelParser:
    """Парсер Excel файлов с полетными данными и SHR телеграммами"""

    @staticmethod
    def parse_excel_date(excel_date) -> Optional[datetime]:
        """Конвертация Excel serial date в datetime"""
        if pd.isna(excel_date):
            return None
        
        # Pandas уже конвертирует datetime объекты
        if isinstance(excel_date, (datetime, pd.Timestamp)):
            if isinstance(excel_date, pd.Timestamp):
                return excel_date.to_pydatetime().replace(hour=0, minute=0, second=0, microsecond=0)
            return excel_date.replace(hour=0, minute=0, second=0, microsecond=0)  # ✅ Правильный отступ
        
        # Теперь этот код достижим!
        if isinstance(excel_date, (int, float)):
            # Excel serial date начинается с 1 января 1900 = 1
            # Но есть баг в Excel с 1900 високосным годом, поэтому вычитаем 2
            try:
                return datetime(1899, 12, 30) + timedelta(days=float(excel_date))
            except:
                return None
        
        if isinstance(excel_date, str):
            formats = ['%d.%m.%Y', '%Y-%m-%d', '%d/%m/%Y', '%d/%m/%y']
            for fmt in formats:
                try:
                    return datetime.strptime(excel_date, fmt)
                except:
                    continue
        
        return None

    @staticmethod
    def parse_flight_data(file_path: str) -> List[Dict[str, Any]]:
        """
        Парсинг Excel файла с SHR телеграммами.
        Обрабатывает оба формата:
        1. Файлы с множественными листами (2024.xlsx)
        2. Файлы с одним листом Result_1 (2025.xlsx)
        """
        from app.services.parser import TelegramParser
        
        all_flights = []
        
        try:
            # Загружаем workbook для проверки структуры
            wb = load_workbook(file_path, read_only=True, data_only=True)
            sheet_names = wb.sheetnames
            logger.info(f"Найдено листов: {len(sheet_names)}: {sheet_names}")
            
            for sheet_name in sheet_names:
                logger.info(f"Обработка листа: {sheet_name}")
                
                # Читаем лист
                df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
                
                # Определяем формат файла
                if df.shape[0] < 2:
                    logger.warning(f"Лист {sheet_name} пустой или слишком мал")
                    continue
                
                # Проверяем первую строку для определения формата
                first_row = df.iloc[0].fillna('').astype(str).str.lower()
                
                # Формат 1: Дата полёта, Сообщение SHR, Сообщение DEP, Сообщение ARR
                if 'дата' in ' '.join(first_row) or 'полёта' in ' '.join(first_row):
                    flights = ExcelParser._parse_telegram_format(df, sheet_name)
                    all_flights.extend(flights)
                
                # Формат 2: Центр ЕС ОрВД, SHR, DEP, ARR
                elif 'центр' in ' '.join(first_row) or 'орвд' in ' '.join(first_row):
                    flights = ExcelParser._parse_center_format(df)
                    all_flights.extend(flights)
                
                # Формат 3: Обычный формат с колонками (старый формат)
                else:
                    flights = ExcelParser._parse_standard_format(df)
                    all_flights.extend(flights)
            
            wb.close()
            logger.info(f"Всего распарсено {len(all_flights)} полетов")
            return all_flights
            
        except Exception as e:
            logger.error(f"Ошибка парсинга Excel файла: {e}", exc_info=True)
            raise

    @staticmethod
    def _parse_telegram_format(df: pd.DataFrame, region_name: str) -> List[Dict[str, Any]]:
        """
        Парсинг формата с телеграммами:
        Row 0: Заголовок периода
        Row 1: Дата полёта | Сообщение SHR | Сообщение DEP | Сообщение ARR
        Row 2+: Данные
        """
        from app.services.parser import TelegramParser
        from app.services.shr_parser import SHRDataParser
        
        flights = []
        
        # Пропускаем первые 2 строки (заголовок периода и названия колонок)
        data_start = 2
        
        for idx in range(data_start, len(df)):
            try:
                row = df.iloc[idx]
                
                # Проверяем, что строка не пустая
                if pd.isna(row[0]):
                    continue
                
                # Парсим дату
                flight_date = ExcelParser.parse_excel_date(row[0])
                if not flight_date:
                    logger.warning(f"Не удалось распарсить дату в строке {idx}: {row[0]}")
                    continue
                
                # Получаем телеграммы
                shr_text = str(row[1]) if not pd.isna(row[1]) else ""
                dep_text = str(row[2]) if len(row) > 2 and not pd.isna(row[2]) else ""
                arr_text = str(row[3]) if len(row) > 3 and not pd.isna(row[3]) else ""
                
                # Создаем объект полета
                flight_data = {
                    'date': flight_date,
                    'region': region_name,
                    'raw_shr': shr_text,
                    'raw_dep': dep_text,
                    'raw_arr': arr_text
                }
                
                # Парсим SHR телеграмму
                if shr_text and 'SHR' in shr_text:
                    parsed_shr = TelegramParser.parse_shr_message(shr_text)
                    flight_data.update({
                        'sid': parsed_shr.get('sid'),
                        'dep_coords': parsed_shr.get('dep_coords'),
                        'arr_coords': parsed_shr.get('arr_coords'),
                        'operator': parsed_shr.get('operator'),
                        'uav_type': parsed_shr.get('uav_type'),
                        'registration': parsed_shr.get('registration'),
                        'altitude': parsed_shr.get('altitude'),
                    })
                
                # Парсим DEP телеграмму
                if dep_text and 'DEP' in dep_text:
                    parsed_dep = TelegramParser.parse_dep_message(dep_text)
                    
                    # Время вылета
                    if parsed_dep.get('atd') and parsed_dep.get('add'):
                        try:
                            date_str = parsed_dep['add']
                            time_str = parsed_dep['atd']
                            
                            year = 2000 + int(date_str[0:2])
                            month = int(date_str[2:4])
                            day = int(date_str[4:6])
                            hour = int(time_str[0:2])
                            minute = int(time_str[2:4])
                            
                            flight_data['dep_time'] = datetime(year, month, day, hour, minute)
                        except Exception as e:
                            logger.debug(f"Не удалось распарсить время DEP: {e}")
                    
                    # Координаты из DEP (если не было в SHR)
                    if not flight_data.get('dep_coords') and parsed_dep.get('adepz'):
                        flight_data['dep_coords'] = parsed_dep['adepz']
                
                # Парсим ARR телеграмму
                if arr_text and 'ARR' in arr_text:
                    parsed_arr = TelegramParser.parse_arr_message(arr_text)
                    
                    # Время прилета
                    if parsed_arr.get('ata') and parsed_arr.get('ada'):
                        try:
                            date_str = parsed_arr['ada']
                            time_str = parsed_arr['ata']
                            
                            year = 2000 + int(date_str[0:2])
                            month = int(date_str[2:4])
                            day = int(date_str[4:6])
                            hour = int(time_str[0:2])
                            minute = int(time_str[2:4])
                            
                            flight_data['arr_time'] = datetime(year, month, day, hour, minute)
                        except Exception as e:
                            logger.debug(f"Не удалось распарсить время ARR: {e}")
                    
                    # Координаты из ARR (если не было в SHR)
                    if not flight_data.get('arr_coords') and parsed_arr.get('adarrz'):
                        flight_data['arr_coords'] = parsed_arr['adarrz']
                
                # Рассчитываем продолжительность
                if flight_data.get('dep_time') and flight_data.get('arr_time'):
                    duration = flight_data['arr_time'] - flight_data['dep_time']
                    # Если отрицательная - прилет на следующий день
                    if duration.total_seconds() < 0:
                        flight_data['arr_time'] += timedelta(days=1)
                        duration = flight_data['arr_time'] - flight_data['dep_time']
                    flight_data['duration_minutes'] = int(duration.total_seconds() / 60)
                
                flights.append(flight_data)
                
            except Exception as e:
                logger.warning(f"Ошибка парсинга строки {idx}: {e}")
                continue
        
        logger.info(f"Из листа {region_name} распарсено {len(flights)} полетов")
        return flights

    @staticmethod
    def _parse_center_format(df: pd.DataFrame) -> List[Dict[str, Any]]:
        """
        Парсинг формата с центрами:
        Row 0: Центр ЕС ОрВД | SHR | DEP | ARR
        Row 1+: Данные
        """
        from app.services.parser import TelegramParser
        
        flights = []
        
        # Пропускаем заголовок
        for idx in range(1, len(df)):
            try:
                row = df.iloc[idx]
                
                # Проверяем, что строка не пустая
                if pd.isna(row[0]):
                    continue
                
                center_name = str(row[0]).strip()
                shr_text = str(row[1]) if not pd.isna(row[1]) else ""
                dep_text = str(row[2]) if len(row) > 2 and not pd.isna(row[2]) else ""
                arr_text = str(row[3]) if len(row) > 3 and not pd.isna(row[3]) else ""
                
                flight_data = {
                    'center_name': center_name,
                    'raw_shr': shr_text,
                    'raw_dep': dep_text,
                    'raw_arr': arr_text
                }
                
                # Парсим телеграммы (аналогично _parse_telegram_format)
                if shr_text and 'SHR' in shr_text:
                    parsed_shr = TelegramParser.parse_shr_message(shr_text)
                    flight_data.update({
                        'sid': parsed_shr.get('sid'),
                        'date': parsed_shr.get('date'),
                        'dep_coords': parsed_shr.get('dep_coords'),
                        'arr_coords': parsed_shr.get('arr_coords'),
                        'operator': parsed_shr.get('operator'),
                        'uav_type': parsed_shr.get('uav_type'),
                        'registration': parsed_shr.get('registration'),
                        'altitude': parsed_shr.get('altitude'),
                    })
                
                if dep_text and 'DEP' in dep_text:
                    parsed_dep = TelegramParser.parse_dep_message(dep_text)
                    if parsed_dep.get('atd') and parsed_dep.get('add'):
                        try:
                            date_str = parsed_dep['add']
                            time_str = parsed_dep['atd']
                            year = 2000 + int(date_str[0:2])
                            month = int(date_str[2:4])
                            day = int(date_str[4:6])
                            hour = int(time_str[0:2])
                            minute = int(time_str[2:4])
                            flight_data['dep_time'] = datetime(year, month, day, hour, minute)
                        except:
                            pass
                
                if arr_text and 'ARR' in arr_text:
                    parsed_arr = TelegramParser.parse_arr_message(arr_text)
                    if parsed_arr.get('ata') and parsed_arr.get('ada'):
                        try:
                            date_str = parsed_arr['ada']
                            time_str = parsed_arr['ata']
                            year = 2000 + int(date_str[0:2])
                            month = int(date_str[2:4])
                            day = int(date_str[4:6])
                            hour = int(time_str[0:2])
                            minute = int(time_str[2:4])
                            flight_data['arr_time'] = datetime(year, month, day, hour, minute)
                        except:
                            pass
                
                # Рассчитываем продолжительность
                if flight_data.get('dep_time') and flight_data.get('arr_time'):
                    duration = flight_data['arr_time'] - flight_data['dep_time']
                    if duration.total_seconds() < 0:
                        flight_data['arr_time'] += timedelta(days=1)
                        duration = flight_data['arr_time'] - flight_data['dep_time']
                    flight_data['duration_minutes'] = int(duration.total_seconds() / 60)
                
                flights.append(flight_data)
                
            except Exception as e:
                logger.warning(f"Ошибка парсинга строки {idx}: {e}")
                continue
        
        return flights

    @staticmethod
    def _parse_standard_format(df: pd.DataFrame) -> List[Dict[str, Any]]:
        """
        Парсинг стандартного формата с колонками:
        Дата | Рейс | Борт | Тип ВС | и т.д.
        (старый формат, если встретится)
        """
        flights = []
        
        # Предполагаем, что первая строка - заголовки
        df.columns = df.iloc[0]
        df = df[1:]
        
        for idx, row in df.iterrows():
            try:
                flight_data = {
                    'date': ExcelParser.parse_excel_date(row.get('Дата')),
                    'aircraft': row.get('Борт'),
                    'aircraft_type': row.get('Тип ВС'),
                    'operator': row.get('Оператор'),
                }
                
                if flight_data.get('date'):
                    flights.append(flight_data)
                    
            except Exception as e:
                logger.warning(f"Ошибка парсинга строки: {e}")
                continue
        
        return flights